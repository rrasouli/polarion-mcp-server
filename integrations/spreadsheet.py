"""
Spreadsheet Integration
Import/export test cases from/to Excel and CSV
"""

import os
import csv
from typing import Dict, Any, List, Optional

try:
    from openpyxl import Workbook, load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class SpreadsheetIntegration:
    """Spreadsheet import/export for test cases"""

    def __init__(self, client):
        self.client = client

    def export_test_cases(
        self,
        query: str,
        output_file: str,
        project_id: str,
        include_test_steps: bool = True,
        format: str = "xlsx"
    ) -> Dict[str, Any]:
        """Export test cases to spreadsheet"""

        # Search for test cases
        search_result = self.client.search_test_cases(
            query=query,
            project_id=project_id,
            limit=1000  # Max limit
        )

        if search_result.get("status") != "success":
            return {
                "status": "failed",
                "error": search_result.get("error", "Search failed")
            }

        test_cases = search_result.get("test_cases", [])

        if not test_cases:
            return {
                "status": "failed",
                "error": "No test cases found matching query"
            }

        # Get full details for each test case
        detailed_cases = []
        for tc in test_cases:
            tc_id = tc["id"].split("/")[-1]  # Extract ID from full path
            details = self.client.get_test_case(
                test_case_id=tc_id,
                project_id=project_id,
                include_test_steps=include_test_steps
            )
            if details.get("status") == "success":
                detailed_cases.append(details)

        # Export based on format
        if format == "xlsx" and OPENPYXL_AVAILABLE:
            return self._export_to_xlsx(detailed_cases, output_file, include_test_steps)
        elif format == "csv":
            return self._export_to_csv(detailed_cases, output_file, include_test_steps)
        else:
            return {
                "status": "failed",
                "error": f"Format {format} not supported or openpyxl not installed"
            }

    def _export_to_xlsx(
        self,
        test_cases: List[Dict],
        output_file: str,
        include_test_steps: bool
    ) -> Dict[str, Any]:
        """Export to Excel format"""

        wb = Workbook()
        ws = wb.active
        ws.title = "Test Cases"

        # Headers
        headers = ["ID", "Title", "Status", "Severity", "Description"]
        if include_test_steps:
            headers.extend(["Test Steps", "Expected Results"])

        ws.append(headers)

        # Data rows
        for tc in test_cases:
            row = [
                tc.get("test_case_id", ""),
                tc.get("title", ""),
                tc.get("status", ""),
                tc.get("severity", ""),
                tc.get("description", "")[:1000]  # Limit description length
            ]

            if include_test_steps:
                # Extract test steps from description or attributes
                row.extend(["", ""])  # Placeholder for now

            ws.append(row)

        wb.save(output_file)

        return {
            "status": "success",
            "message": f"Exported {len(test_cases)} test cases to {output_file}",
            "output_file": output_file,
            "test_cases_count": len(test_cases)
        }

    def _export_to_csv(
        self,
        test_cases: List[Dict],
        output_file: str,
        include_test_steps: bool
    ) -> Dict[str, Any]:
        """Export to CSV format"""

        with open(output_file, 'w', newline='', encoding='utf-8') as csvfile:
            headers = ["ID", "Title", "Status", "Severity", "Description"]
            if include_test_steps:
                headers.extend(["Test Steps", "Expected Results"])

            writer = csv.DictWriter(csvfile, fieldnames=headers)
            writer.writeheader()

            for tc in test_cases:
                row = {
                    "ID": tc.get("test_case_id", ""),
                    "Title": tc.get("title", ""),
                    "Status": tc.get("status", ""),
                    "Severity": tc.get("severity", ""),
                    "Description": tc.get("description", "")[:1000]
                }

                if include_test_steps:
                    row["Test Steps"] = ""
                    row["Expected Results"] = ""

                writer.writerow(row)

        return {
            "status": "success",
            "message": f"Exported {len(test_cases)} test cases to {output_file}",
            "output_file": output_file,
            "test_cases_count": len(test_cases)
        }

    def import_test_cases(
        self,
        spreadsheet_file: str,
        project_id: str,
        update_existing: bool = False
    ) -> Dict[str, Any]:
        """Import test cases from spreadsheet"""

        if not os.path.exists(spreadsheet_file):
            return {
                "status": "failed",
                "error": f"File not found: {spreadsheet_file}"
            }

        # Determine format
        if spreadsheet_file.endswith(".xlsx") and OPENPYXL_AVAILABLE:
            rows = self._read_xlsx(spreadsheet_file)
        elif spreadsheet_file.endswith(".csv"):
            rows = self._read_csv(spreadsheet_file)
        else:
            return {
                "status": "failed",
                "error": "Unsupported file format or openpyxl not installed"
            }

        results = {
            "total": len(rows),
            "created": 0,
            "updated": 0,
            "errors": []
        }

        for row in rows:
            title = row.get("Title", "").strip()
            description = row.get("Description", "").strip()
            severity = row.get("Severity", "should_have").strip()
            status = row.get("Status", "draft").strip()

            if not title:
                results["errors"].append("Skipped row with no title")
                continue

            # Create test case
            create_result = self.client.create_test_case(
                title=title,
                description=description,
                project_id=project_id,
                severity=severity,
                status=status
            )

            if create_result.get("status") == "success":
                results["created"] += 1

                # Add test steps if provided
                test_steps_text = row.get("Test Steps", "").strip()
                expected_results_text = row.get("Expected Results", "").strip()

                if test_steps_text and expected_results_text:
                    # Parse newline-separated steps
                    steps = test_steps_text.split("\n")
                    expected = expected_results_text.split("\n")

                    test_steps = []
                    for i, step in enumerate(steps):
                        test_steps.append({
                            "step": step.strip(),
                            "expectedResult": expected[i].strip() if i < len(expected) else ""
                        })

                    if test_steps:
                        test_case_id = create_result.get("test_case_id", "").split("/")[-1]
                        self.client.add_test_steps(
                            test_case_id=test_case_id,
                            test_steps=test_steps,
                            project_id=project_id
                        )
            else:
                results["errors"].append(
                    f"Failed to create '{title}': {create_result.get('error')}"
                )

        return {
            "status": "success",
            "message": f"Imported {results['created']} test cases",
            "spreadsheet_file": spreadsheet_file,
            "statistics": results
        }

    def _read_xlsx(self, file_path: str) -> List[Dict[str, str]]:
        """Read Excel file"""
        wb = load_workbook(file_path)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        rows = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = {}
            for i, value in enumerate(row):
                if i < len(headers):
                    row_dict[headers[i]] = str(value) if value else ""
            rows.append(row_dict)

        return rows

    def _read_csv(self, file_path: str) -> List[Dict[str, str]]:
        """Read CSV file"""
        rows = []
        with open(file_path, 'r', encoding='utf-8') as csvfile:
            reader = csv.DictReader(csvfile)
            for row in reader:
                rows.append(row)
        return rows
